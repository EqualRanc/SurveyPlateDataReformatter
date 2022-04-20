# -*- coding: utf-8 -*-
"""
Created on Wed Apr 13 13:46:49 2022

This code is intended to take .xml files from Echo plate surveys and reinterpret the volume data and rearrange into a 384W-plate formatted view.

@author: Chris
"""

import xlwings as xw
from xml.etree import ElementTree as ET
import pandas as pd
import os
import PySimpleGUI as sg
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Color #can also add PatternFill, Font, Border
#from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule #Can also add CellIsRule, FormulaRule
#from openpyxl.styles import colors


threeeightfourheaders = ['','1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24']
letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
fletters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF']
numcol = list(range(1,25))
fnumcol = list(range(1,49))

def platemaker(root):
    pp = []
    newrow = list(range(24,384,24))
    fnewrow = list(range(48,1536,48))
    if root.attrib['totalWells'] == '384':
        row = []
        for x in range(384):
            well = root[x].attrib
            if x in newrow:
                pp.append(row)
                row = []
            row.append(well['vl'])
        pp.append(row)
    if root.attrib['totalWells'] == '1536':
        row = []
        for x in range(1536):
            well = root[x].attrib
            if x in fnewrow:
                pp.append(row)
                row = []
            row.append(well['vl'])
        pp.append(row)
    return pp

def excelsummary(foldername):
    path = os.path.expanduser(foldername)
    if not os.path.exists(path):
        return path
    ls = os.listdir(path)
    ls2 = []
    for temp in ls:
        if temp[-4:] == '.xml':
            if temp[:0] != '~':
                if temp[-5:] != '.xlsx':
                    ls2.append(temp)
    summary = xw.Book()
    ssheet = summary.sheets[0]
    rowinc = 0
    platestrings = []
    for file in ls2:
        filename = path + '/' + file
        xml = ET.parse(filename)
        root = xml.getroot()
        bc = str(root.attrib['barcode'])
        plate = platemaker(root)
        
        rowbc = 1 + rowinc
        rowdf = 2 + rowinc
        rowlet = 3 + rowinc
        posbc = 'A' + str(rowbc)
        posdf = 'B' + str(rowdf)
        poslet = 'B' + str(rowlet)
        
        if root.attrib['totalWells'] == '1536':
            rowinc = rowinc + 36
            df = pd.DataFrame(plate, columns = fnumcol)
            ssheet.range(posbc).value = bc
            ssheet.range(posdf).value = df
            ssheet.range(poslet).options(transpose=True).value = fletters
            stop = rowlet + 31
            posc = 'C' + str(rowlet) + ':AX' + str(stop)
            platestrings.append(posc)
#            ssheet.conditional_format(posc, {'type':'2_color_scale', 'max_color':'#008000', 'min_color':'#FFFFFF'})
        elif root.attrib['totalWells'] == '384':
            rowinc = rowinc + 20
            df = pd.DataFrame(plate, columns = numcol)
            ssheet.range(posbc).value = bc
            ssheet.range(posdf).value = df
            ssheet.range(poslet).options(transpose=True).value = letters
            stop = rowlet + 15
            posc = 'C' + str(rowlet) + ':Z' + str(stop)
            platestrings.append(posc)
#            ssheet.conditional_format(posc, {'type':'2_color_scale', 'max_color':'#008000', 'min_color':'#FFFFFF'})
            
    return summary, platestrings

def excelformat(formatfile,platestr):
    formatpath = formatfile
    wb = load_workbook(formatpath)
    ws = wb.active
    for cclr in platestr:
        ws.conditional_formatting.add(cclr, ColorScaleRule(start_type='min', start_color=Color(index=2), end_type='max', end_color=Color(index=3)))
        #Green is 3, White is 1, Red is 2, Pale Yellow is 26
    wb.save(formatfile)



#Setup for the singple-point data processing main interface
dataentries = [
     [sg.Text("Plate Survey Summarizer",font='Any 18')],
     [sg.Frame('Browse to exported plate survey files:', [[sg.Input(key='-raw-'), sg.FolderBrowse(target='-raw-')]])],
     [sg.Frame("Please give your summary a name:", [[sg.Input(key='-name-')]])],
     [sg.Submit(), sg.Cancel()]
]


datastatus = [
    [sg.Text('Status:', size=[20,1])],
     [sg.Multiline(key='datastatus',autoscroll=True,size=(30,20))],
]

datalayout = [
    [
     sg.Column(dataentries),
     sg.VSeperator(),
     sg.Column(datastatus)
     ]
]

#Creates the theme
sg.theme('Dark Teal')

#Creates the window
window = sg.Window('Carmot Plate Survey Summary Tool', datalayout, no_titlebar=False, alpha_channel=.9, grab_anywhere=True)


#Create event loop to enable user inputs
while True:
    event, values = window.read()
    if event in (sg.WINDOW_CLOSED, "Cancel"):
        break
    
    #Processing steps for renaming assay plates tab
    elif event == "Submit":
        try:
            foldername = values['-raw-']
            name = values['-name-']
            output, platestrings = excelsummary(foldername)
            outfile = str(values['-raw-']) + '/' + str(datetime.date.today().isoformat()) + '_' + '%s' % str(values['-name-']) + '_' + 'Survey_Summary.xlsx'
            output.save(path=outfile)
            output.close()
            excelformat(outfile,platestrings)
            window['datastatus'].print("Processing of survey xml files complete.")
        except:
            window['datastatus'].print("Unable to process survey xml files.")
window.close()